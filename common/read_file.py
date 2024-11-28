# coding:utf-8
import xlrd, csv
from common import config_manage
import openpyxl
import os
from datetime import datetime


def get_data_excel(excelfile_Name, sheetName='Sheet1'):
    excelfile_path = config_manage.TESTDATA_PATH + excelfile_Name + ".xlsx"

    # 检查文件是否存在
    if not os.path.exists(excelfile_path):
        raise FileNotFoundError(f"文件 {excelfile_path} 不存在")

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(excelfile_path)

    # 获取指定的工作表
    sheet = workbook[sheetName]

    # 获取第一行作为 key 值
    keys = [cell.value for cell in sheet[1]]

    # 获取总行数
    rowNum = sheet.max_row

    # 获取总列数
    colNum = sheet.max_column

    if rowNum <= 1:
        print("总行数小于1")
    else:
        r = []
        for j in range(2, rowNum + 1):  # 从第二行开始读取数据
            s = {}
            for x in range(colNum):
                cell = sheet.cell(row=j, column=x + 1).value
                if isinstance(cell, float) and cell.is_integer():  # 如果是整型
                    cell = int(cell)
                elif isinstance(cell, datetime):  # 如果是日期时间
                    cell = cell.strftime('%Y/%d/%m %H:%M:%S')
                elif isinstance(cell, bool):  # 如果是布尔值
                    cell = cell
                s[keys[x]] = cell
            r.append(s)
        return r

def get_data_csv(csvfile_Name):
    csvfile_path = config_manage.TESTDATA_PATH + csvfile_Name + '.csv'
    with open(csvfile_path, "r", encoding="gbk") as f:
        reader = csv.DictReader(f)
        column = [row for row in reader]
    return column


if __name__ == '__main__':
    filepath = "ddtt"
    print(get_data_excel(filepath))
